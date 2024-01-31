from tkinter import *
import ttkbootstrap as tb
import os
import pandas as pd
import requests
from requests.auth import HTTPDigestAuth
import sys
from tkinter import messagebox
from openpyxl import load_workbook
from datetime import datetime, timedelta,date
import calendar
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Border, Side
import locale

def laplace(username, password, filename, time, month, fileyear, downloads_folder, workbooks_folder,params,year,address,cover_name,FIT,PV,formatdate):
    
    def write_date_on_column(current_date, end_date, current_cell,initial_row,final_row):
        while current_date <= end_date:
            current_cell.value = current_date
            current_cell.number_format = 'm"月"d"日("aaa")"'
            current_date += timedelta(days=1)
            current_cell = sheet.cell(row=current_cell.row + 1, column=current_cell.column)
            for row in range(initial_row, final_row):
                for col in sheet.iter_cols(min_row=row, max_row=row):
                    for cell in col:
                        if isinstance(cell, openpyxl.cell.cell.MergedCell):
                            continue
                        else:
                            cell.value = None
    
    api_url = "https://services.energymntr.com/megasolar/{}/services/api/download/monthly.php".format(username)
    try:
        response = requests.get(api_url, params=params, auth=HTTPDigestAuth(
            username, password), stream=True)
        output_path_csv = os.path.join(
            downloads_folder, "{}{}月.csv".format(username, month))
        input_path_workbook = os.path.join(
            workbooks_folder, "保守点検報告書_ベース.xlsx")
        if response.status_code == 200:
            with open(output_path_csv, "wb") as csv_file:
                for chunk in response.iter_content(chunk_size=1024):
                    if chunk:
                        csv_file.write(chunk)
            data = pd.read_csv(output_path_csv, encoding='cp932')
            output_path_excel = os.path.join(
                downloads_folder, "{}{}月.xlsx".format(filename, fileyear))
            data.to_excel(output_path_excel, index=False)
            wb3 = load_workbook(output_path_excel)
            wb1 = load_workbook(input_path_workbook)
            Prevalue=wb3.active
            Prevalue= wb3.create_sheet('PreValue')
            Prevalue['A1'] = '日射量(kWh/m2)'
            Prevalue['B1'] = '売電電力量(kWh)'
            wb3.save(output_path_excel)
            wb3 = load_workbook(output_path_excel)
            
            df = pd.read_excel(output_path_excel, sheet_name='Sheet1')
            source_column_name1 = '売電電力量(kWh)'
            source_column_name2 = '日射量(kWh/m2)'
            source_column_name3 = '日射量2(kWh/m2)'
            column_names = df.columns
            # Load the data from the destination sheet into a dataframe
            destination_sheet_name = 'PreValue'
            destination_column_name1 = '売電電力量(kWh)'  # Change this to the desired destination column name
            destination_column_name2 = '日射量(kWh/m2)'
            df_destination = pd.read_excel(output_path_excel, sheet_name=destination_sheet_name)
            
            if username == "UAP0170823":
                df_destination[destination_column_name1] = df[source_column_name1]
                df_destination[destination_column_name2] = df[source_column_name3]
                
                # Save the modified dataframe back to the Excel file
                with pd.ExcelWriter(output_path_excel, engine='xlsxwriter') as writer:
                    df.to_excel(writer, sheet_name='Sheet1', index=False)
                    df_destination.to_excel(writer, sheet_name=destination_sheet_name, index=False)
            else:
                df_destination[destination_column_name1] = df[source_column_name1]
                df_destination[destination_column_name2] = df[source_column_name2]
                
                # Save the modified dataframe back to the Excel file
                with pd.ExcelWriter(output_path_excel, engine='xlsxwriter') as writer:
                    df.to_excel(writer, sheet_name='Sheet1', index=False)
                    df_destination.to_excel(writer, sheet_name=destination_sheet_name, index=False)
                    
            wb3 = load_workbook(output_path_excel)
            for row in wb3['Sheet1'].iter_rows(values_only=True):
                wb1['ログ'].append(row)
            for row in wb3['PreValue'].iter_rows(values_only=True):
                wb1['Data'].append(row)
            
            rangeselected1 = []
            rangeselected2 = []
            sheet1 = wb1['Data']
            sheet2 = wb1['発電実績']
            
            # copy data of KwH and KWh/m2 of Data sheet to 発電実績 sheet
            for i in range(1,33,1):
                rangeselected1.append(sheet1.cell(row = i,column=1).value)
            
            for i in range(17,50):
                if i-17 < len(rangeselected1):
                    sheet2.cell(row=i,column=25).value = rangeselected1[i-17]
                else:
                    sheet2.cell(row=i,column=25).value = None
            for i in range(1,33,1):
                rangeselected2.append(sheet1.cell(row = i,column=2).value)
            
            for i in range(17,50):
                if i-17 < len(rangeselected2):
                    sheet2.cell(row=i,column=26).value = rangeselected2[i-17]
                else:
                    sheet2.cell(row=i,column=26).value = None
                    
            # Save the modified workbook
            output_path_workbook = os.path.join(
                downloads_folder, "{}{}.xlsx".format(filename, fileyear))
            wb1.save(output_path_workbook)
            df = pd.read_excel(output_path_workbook,sheet_name='Data')
            numbers_of_rows = len(df[source_column_name2])
            sheet_names =['表紙 ','運転監視','発電実績','損失量計算']
            workbook = load_workbook(output_path_workbook)
            
            try:
                for sheet_name in sheet_names:
                    # Check if the sheet exists in the workbook
                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        if sheet_name == '表紙 ':
                            FS = workbook.active
                            FS = workbook[sheet_name]
                            year = int(year)
                            month = int(month)
                            first_day = 1
                            _ , Last_date = calendar.monthrange(year,month)
                            date_object = datetime.strptime('{}-{}-{}'.format(month, first_day, year),'%m-%d-%Y')
                            last_object = datetime.strptime('{}-{}-{}'.format(month,Last_date,year),'%m-%d-%Y')
                            today = date.today()
                            datecell = FS['J15']
                            datecell.value = date_object
                            datecell.number_format = 'yyyy"年"m"月"d"日("aaa")"'
                            datecell = FS['J16']
                            datecell.value = last_object
                            datecell.number_format = 'd"日("aaa")"'
                            datecell = FS['C18']
                            datecell.value = '="運転監視："&TEXT(J15,"yyyy年m月ｄ日")&TEXT(J15,"(aaa)")&" ~ "&TEXT(J16,"ｄ日")&TEXT(J16,"(aaa)")'
                            datecell = FS['B23']
                            datecell.value = '=YEAR(J15)&"年"&MONTH(J15)&"月分の運転監視業務の結果をご報告いたします。"'
                            datacell = FS['F5']
                            datacell.value = today
                            datacell.number_format = 'yyyy"年"m"月"d"日"'
                            
                            #excel formula
                            # ="運転監視："&TEXT(J15,"yyyy年m月ｄ日")&TEXT(J15,"(aaa)")&" ~ "&TEXT(J16,"ｄ日")&TEXT(J16,"(aaa)")
                            # =YEAR(J15)&"年"&MONTH(J15)&"月分の運転監視業務の結果をご報告いたします。"
                            FS['A3'] = cover_name
                            FS['C16'] = filename
                            FS['C17'] = address
                        elif sheet_name == '運転監視':
                            start_date_str = '{}-{}-{}'.format(month, first_day, year)
                            start_date = datetime.strptime(start_date_str, "%m-%d-%Y")
                            end_date_str = '{}-{}-{}'.format(month, numbers_of_rows, year)
                            end_date = datetime.strptime(end_date_str, "%m-%d-%Y")

                            final_row = 44
                            day_difference = 31 - numbers_of_rows
                            initial_row =  final_row - day_difference
                            # print(initial_row)
                            start_cell = 'B13'
                            current_date = start_date
                            current_cell = sheet[start_cell]
                            write_date_on_column(current_date, end_date, current_cell,initial_row,final_row)
                            SS = workbook.active
                            SS = workbook[sheet_name]
                            num_rows_to_delete = final_row - initial_row
                            SS.delete_rows(initial_row,num_rows_to_delete)
                            # if num_rows_to_delete != 0:
                            top =  Border( right=Side(border_style='thin'), 
                                                    top=Side(border_style='thick'), 
                                                    bottom=Side(border_style='thin'))
                            lside =  Border( left=Side(border_style='thick'), 
                                                    top=Side(border_style='thin'),
                                                    right=Side(border_style='thin'), 
                                                    bottom=Side(border_style='thin'))
                            rside =  Border( left=Side(border_style='thin'), 
                                                    top=Side(border_style='thin'),
                                                    right=Side(border_style='thick'), 
                                                    bottom=Side(border_style='thin'))
                            buttom =  Border( right=Side(border_style='thin'), 
                                                    top=Side(border_style='thin'), 
                                                    bottom=Side(border_style='thick'))
                            for col in SS.iter_cols(min_col=2, max_col=8):
                                for cell in col:
                                    if cell.row == 12:
                                        cell.border = top
                            
                            for row in SS.iter_rows(min_row=12, max_row=initial_row-2):
                                for cell in row:
                                    if cell.column == 2:
                                        cell.border = lside
                            for row in SS.iter_rows(min_row=12, max_row=initial_row-2):
                                for cell in row:
                                    if cell.column == 8:
                                        cell.border = rside
                                        
                            for col in SS.iter_cols(min_col=2, max_col=8):
                                for cell in col:
                                    if cell.row == initial_row - 1:
                                        cell.border = buttom
                            left =  Border( left=Side(border_style='thick'),
                                           bottom=Side(border_style='thick'),
                                           right=Side(border_style='thin'))
                            topfirst =  Border( left=Side(border_style='thick'),
                                           bottom=Side(border_style='thin'),
                                           right=Side(border_style='thin'),
                                           top=Side(border_style='thick'))
                            right =  Border( left=Side(border_style='thin'),
                                           bottom=Side(border_style='thick'),
                                           right=Side(border_style='thick'))
                            leftfirst =  Border( left=Side(border_style='thin'),
                                           bottom=Side(border_style='thin'),
                                           right=Side(border_style='thick'),
                                           top=Side(border_style='thick'))
                            SS['B'+str(initial_row-1)].border = left
                            SS['B12'].border = topfirst
                            SS['H12'].border = leftfirst
                            SS['H'+str(initial_row-1)].border = right
                        
                        elif sheet_name == '発電実績':
                            
                            start_date_str = '{}-{}-{}'.format(month, first_day, year)
                            start_date = datetime.strptime(start_date_str, "%m-%d-%Y")
                            end_date_str = '{}-{}-{}'.format(month, numbers_of_rows, year)
                            end_date = datetime.strptime(end_date_str, "%m-%d-%Y")

                            final_row = 49
                            day_difference = 31 - numbers_of_rows
                            initial_row =  final_row - day_difference
                            cfinalvalue = initial_row-1
                            # Specify the cell in which to start writing dates
                            start_cell = 'A18'
                            current_date = start_date
                            current_cell = sheet[start_cell]
                            write_date_on_column(current_date, end_date, current_cell,initial_row,final_row)
                            TS = workbook.active
                            TS = workbook[sheet_name]
                            # Calculate the number of rows to delete
                            num_rows_to_delete = final_row - initial_row
                            if num_rows_to_delete !=0:
                                TS.delete_rows(initial_row,num_rows_to_delete)
                            
                            TS['C9'] = formatdate
                            TS['E9'] = FIT
                            TS['F9'] = PV
                            TS['I9'] = '無'
                            datacell = TS['B'+str(initial_row)]
                            datacell.value = f'=SUM(B18:B{cfinalvalue})'
                            datacell = TS['C'+str(initial_row)]
                            datacell.value = f'=SUM(C18:C{cfinalvalue})'
                            datacell = TS['D'+str(initial_row)]
                            datacell.value = f'=SUM(D18:D{cfinalvalue})'
                            datacell = TS['F'+str(initial_row)]
                            datacell.value = f'=SUM(F18:F{cfinalvalue})'
                            datacell = TS['G'+str(initial_row)]
                            datacell.value = f'=SUM(G18:G{cfinalvalue})'
                            datacell = TS['I'+str(initial_row)]
                            datacell.value = f'=SUM(I18:I{cfinalvalue})'
                            datacell = TS['J'+str(initial_row)]
                            datacell.value = f'=SUM(J18:J{cfinalvalue})'
                            datacell = TS['P'+str(initial_row)]
                            datacell.value = f'=MOD(SUM(P18:P{cfinalvalue}),60)'
                            datacell = TS['O'+str(initial_row)]
                            datacell.value = f'=SUM(O18:O{cfinalvalue}) + INT(SUM(P18:P{cfinalvalue})/60)'      
                            datacell = TS['G9']
                            datacell.value = f'=AVERAGE(E18:E{cfinalvalue})'
                            datacell = TS['H9']
                            datacell.value = f'=I{cfinalvalue+1}/F{cfinalvalue+1}'    
                            start_row = 18
                            end_row = cfinalvalue

                            # Loop over the range of rows
                            # for row in range(start_row, end_row + 1):
                            #     # Get the cell value from column 2
                            #     src_cell = TS.cell(row=row, column=20)

                            #     # Set the value to the corresponding cell in column 5
                            #     dest_cell = TS.cell(row=row, column=2)
                            #     dest_cell.value = src_cell.value
                            #     src_cell.value = None
                            # for row in range(start_row, end_row + 1):
                            #     # Get the cell value from column 2
                            #     src_cell = TS.cell(row=row, column=21)

                            #     # Set the value to the corresponding cell in column 5
                            #     dest_cell = TS.cell(row=row, column=3)
                            #     dest_cell.value = src_cell.value
                            #     src_cell.value = None
                            # TS['T17'].value = None
                            # TS['U17'].value = None
                    
                            
                        elif sheet_name == '損失量計算':
                            start_date_str = '{}-{}-{}'.format(month, first_day, year)
                            start_date = datetime.strptime(start_date_str, "%m-%d-%Y")
                            end_date_str = '{}-{}-{}'.format(month, numbers_of_rows, year)
                            end_date = datetime.strptime(end_date_str, "%m-%d-%Y")
                            final_row = 34
                            day_difference = 31 - numbers_of_rows
                            initial_row =  final_row - day_difference
                            finalvalue = initial_row-1
                            # Specify the cell in which to start writing dates
                            start_cell = 'A3'  
                            current_date = start_date
                            current_cell = sheet[start_cell]
                            write_date_on_column(current_date, end_date, current_cell,initial_row,final_row)
                            TS = workbook.active
                            TS = workbook[sheet_name]
                            # Calculate the number of rows to delete
                            num_rows_to_delete = final_row - initial_row
                            TS['K3'] = PV
                            if num_rows_to_delete !=0:
                                TS.delete_rows(initial_row,num_rows_to_delete)
                                     
                    else:
                        print(f"Sheet '{sheet_name}' not found in the workbook.")
                    
            finally:
                workbook.remove(workbook['Data'])
                workbook.save(output_path_workbook)
                os.remove(output_path_csv) 
                os.remove(output_path_excel)
                wb = load_workbook(output_path_workbook)
                ws = wb['発電実績']
                values = Reference(ws, min_col=25, min_row=17, max_col=25, max_row=cfinalvalue)
                values2 = Reference(ws, min_col=26, min_row=17, max_col=26, max_row=cfinalvalue) 
                barchart = BarChart()
                barchart.add_data(values2,titles_from_data=True)
                barchart.x_axis.title = '日付'
                # barchart.y_axis.scaling.min = 0
                # barchart.y_axis.scaling.max = 8000
                # barchart.y_axis.minorGridlines = None
                # barchart.y_axis.majorGridlines = None
                barchart.height = 12
                barchart.width = 29
                barchart.x_axis.number_format = 'd"日"'
                barchart.y_axis.title = '発電電力量[kWh]'
                barchart.series[0].graphicalProperties.solidFill = '00b050'
                # Add the legend layout here
                barchart.legend.position = 'b'
                linechart = LineChart()
                linechart.add_data(values, titles_from_data=True)
                linechart.y_axis.axId = 200
                linechart.y_axis.minorGridlines = None
                linechart.y_axis.majorGridlines = None
                linechart.y_axis.scaling.min = 0
                linechart.y_axis.scaling.max = 6
                linechart.y_axis.title = '日射量[kWh/m2]'
                linechart.y_axis.crosses = "max"
                linechart.series[0].graphicalProperties.line.solidFill = 'fcd221'
                barchart += linechart
                barchart.title = "{}   発電電力量の実績".format(filename)                
                ws.add_chart(barchart, 'A51') 
                wb.save(output_path_workbook)
    except requests.exceptions.RequestException as e:
        print(f"Error: {str(e)}")                        
                        
def download():
    # selected_year = year_var.get()
    # selected_month = month_var.get()
    # print("{}:{}".format(selected_year,selected_month))
    # locale.setlocale(locale.LC_TIME, '')

    # # Get the date format
    # date_format = locale.nl_langinfo(locale.D_FMT)
    # print('Date format: ', date_format)
    selected_date = mydate.entry.get()
    selected_date = selected_date
    
    start_date = datetime.strptime(selected_date, "%Y-%m-%d")
    # start_date = datetime.strptime(selected_date, "%m/%d/%Y")
    
    formatted_date = start_date.strftime("%Y-%m-%d")
    print(formatted_date)
    print("Download initiated")
    year,month,date= formatted_date.split('-')
    time = year + month
    fileyear = time[-4:]
    def get_downloads_folder():
        current_working_directory = os.getcwd()
        print(current_working_directory)

        # home_dir = os.path.expanduser("~")
        # downloads_folder = os.path.join(home_dir, "Downloads")
        # os.makedirs(downloads_folder, exist_ok=True)
        return current_working_directory

    def get_workbook_folder():
        current_working_directory = os.getcwd()
        # home_dir = os.path.expanduser("~")
        # workbooks_folder = os.path.join(home_dir, "Documents\\excel")
        path = os.path.join( current_working_directory,'保守点検報告書_ベース.xlsx')
        isFile = os.path.isfile(path)
        if isFile == False:
            print("there is no such files")
            messagebox.showerror("エラー", "エクセルファイルが見つかりません。{}パスのベースエクセルシートを確認してください。".format(current_working_directory))
        else:
            print("file found on path {}".format(path))
        # os.makedirs(workbooks_folder, exist_ok=True)
        return current_working_directory

    downloads_folder = get_downloads_folder()
    workbooks_folder = get_workbook_folder()
    counter1 = 0
    # counter2 = 0
    # users = [
    #     {"device_id": "CDA3614231168779955966051050305660734614600","filename":"KIC日高_電力量実績_"},
    #     {"device_id": "CDA1234823544587112915614506550158738364790","filename":"KIC厚木_電力量実績_"},
    #     {"device_id": "CDA3975154446986497297068957620540553993780","filename":"KIC越谷_電力量実績_"},
    #     {"device_id": "CDA4995386155929238379015114546939378371328","filename":"岩根中学校_電力量実績_"},
    #     {"device_id": "CDA1836290764883644058616167299722898321065","filename":"太田中学校_電力量実績_"},
    #     {"device_id": "CDA8626825421151183593093310386921157478810","filename":"清見台小学校_電力量実績_"},
    #     {"device_id": "CDA1229397212628199228294670587600616668420","filename":"富来田中学校_電力量実績_"},
    #     {"device_id": "CDA1066501134050441480013516012951809329362","filename":"畑沢小学校_電力量実績_"},
    #     {"device_id": "CDA1524381332789749178434543690444562842710","filename":"第二中学校_電力量実績_"},
    #     {"device_id": "CDA1433063410857518239315949293158629193154","filename":"請西小学校_電力量実績_"}
    # ]
    params = {
        "groupid": "1",
        "time": time,
        "data": "measuringdata",
        "format": "csv",
        "type": "pcs"
    }
    users1= [
        {"username": "XGM0182319", "password": "MUcjx4kMvCxeMo9","filename":"広野ソーラーパーク","address":"福島県双葉郡広野町折木字東下4-18、外7筆","cover_name":"合同会社SS福島広野","report_name":"保守点検報告書","FIT":'=40',"PV":'=0.29*9372',"opening_date":"12-12-2019"},
        {"username": "CBU0173382", "password": "aiqh9gMgrfLvUgz","filename":"田方郡函南町太陽光発電所","address":"静岡県田方郡函南町軽井沢字浜井場295-8　他10筆","cover_name":"リニューアブル・ジャパン株式会社","report_name":"運転監視報告書","FIT":'=40',"PV":'=312*0.275+7392*0.245',"opening_date":"9-26-2018"},
        # {"username": "FOO0197257", "password": "rdR9YWWNeumoiEL","filename":"千歳新川低圧_運転監視報告書_"},
        # {"username": "MUI020F591", "password": "zr9qjgpwdtW4tRe","filename":"大郷町小学校_電力量実績_"},
        # {"username": "XVK020F592", "password": "kWiEAKcpWc9kfoJ","filename":"大郷町中学校_電力量実績_"},
        # {"username": "SXR020F590", "password": "zzxf4ge3LmKKRuz","filename":"大郷町文化会館_電力量実績_"},
        # {"username": "IQH020C788", "password": "zmeEuMcL4zuYjt4","filename":"大郷町役場_電力量実績_"},
        {"username": "UAP0170823", "password": "AhKC9WrRETcfnTt","filename":"岩手一関ソーラーパーク","address":"岩手県一関市花泉町花泉字大又南沢３－５他地内","cover_name":"合同会社SS岩手一関","report_name":"保守点検報告書","FIT":'=36',"PV":'=0.275*8904',"opening_date":"12-25-2020"},
        {"username": "LLH0162908", "password": "3XC7NhwTsxiRydz","filename":"御田神辺池ソーラー発電所","address":"香川県さぬき市寒川町石田東字御田神辺甲1539","cover_name":"合同会社香川水上ソーラー第二","report_name":"運転監視報告書","FIT":'=32',"PV":'=0.295*5175',"opening_date":"9-8-2017"}
    ]
    for user in users1:
        username = user["username"]
        password = user["password"]
        filename = user["filename"]
        address=user["address"]
        address = str(address)
        cover_name=user["cover_name"]
        cover_name=str(cover_name)
        FIT = user['FIT']
        PV = user['PV']
        opening_date = user['opening_date']
        opening_date = datetime.strptime(opening_date, "%m-%d-%Y")
        # date_object = datetime.strptime(selected_date, "%m/%d/%Y")
        OPD = opening_date.strftime("%Y-%m-%d")
        oyear,omonth,odate= OPD.split('-')
        date_object = datetime.strptime('{}-{}-{}'.format(omonth, odate, oyear),'%m-%d-%Y')
        formatdate = date_object.strftime('%Y年%m月%d日')
        laplace(username, password, filename, time, month, fileyear, downloads_folder, workbooks_folder,params,year,address,cover_name,FIT,PV,formatdate) 
        counter1 += 1
        if counter1 == 4:
            messagebox.showinfo("ダウンロードステータス", "ダウンロード完了")
        # print(f"{counter1} iterations complete in users1")           
    # for user in users:
    #     device_id = user["device_id"]
    #     filename = user["filename"]
    #     SS(device_id,filename,formatted_date,downloads_folder,workbooks_folder,month)
    #     counter2 += 1
    #     # print(f"{counter2} iterations complete in users2")
    #     if counter2 == 10:
    #         messagebox.showinfo("ダウンロードステータス", "ダウンロード完了")
def open_date_picker():
    top = tb.Toplevel(root)
    top.title("日付を選ぶ")
    top.geometry('400x400')
    # Create a frame to enclose the login elements
    container = tb.Frame(top, padding=20, borderwidth=4, relief="solid", width=300, height=250)
    container.place(relx=0.5, rely=0.5, anchor=CENTER)
    global mydate
    mydate = tb.DateEntry(container, bootstyle="danger", width=20,dateformat = r"%Y-%m-%d")  # Increase box size
    mydate.pack(pady=50)
    download_button = tb.Button(container, text="ダウンロード", command=download)
    download_button.pack()
    top.protocol("WM_DELETE_WINDOW", sys.exit)
    # top = tb.Toplevel(root)
    # top.title("年と月を選ぶ")
    # top.geometry('400x400')
    # container = tb.Frame(top, padding=20, borderwidth=4, relief="solid", width=300, height=250)
    # container.place(relx=0.5, rely=0.5, anchor=CENTER)

    # # Create a list of years and months
    # years = [str(year) for year in range(2000, 2035)]
    # months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]

    # # Create dropdown menus for year and month
    # global year_var, month_var
    # year_var = tb.StringVar(top)
    # month_var = tb.StringVar(top)
    # year_menu = tb.OptionMenu(container, year_var, *years)
    # month_menu = tb.OptionMenu(container, month_var, *months)    
    # year_menu.pack(pady=20)
    # month_menu.pack(pady=20)

    # download_button = tb.Button(container, text="ダウンロード", command=download)
    # download_button.pack()

    # top.protocol("WM_DELETE_WINDOW", sys.exit)
def login():
    #default username and password
    default_username = "sems2024"
    default_password = "Sm@rtsol@r"
    username = username_entry.get()
    password = password_entry.get()
    if username == default_username and password == default_password:
        root.withdraw() 
        open_date_picker()
    else:
        messagebox.showerror("エラー", "無効なユーザー名/パスワード")
root = tb.Window(themename="superhero")
root.title("ログインフォーム")
root.geometry('400x400')
style = tb.Style()
style.configure('TEntry', font=('Helvetica', 14))
style.configure('TButton', font=('Helvetica', 14))
style.configure('TLabel', font=('Helvetica', 14))

# Create a frame to enclose the login elements
login_container = tb.Frame(root, padding=20, borderwidth=4, relief="solid", width=300, height=250)
login_container.place(relx=0.5, rely=0.5, anchor=CENTER)

username_label = tb.Label(login_container, text="ユーザー名")
username_label.grid(row=0, column=0, padx=10, pady=10, columnspan=2)
username_entry = tb.Entry(login_container, width=20) 
username_entry.grid(row=1, column=0, padx=10, pady=10, columnspan=2)

password_label = tb.Label(login_container, text="パスワード")
password_label.grid(row=2, column=0, padx=10, pady=10, columnspan=2)
password_entry = tb.Entry(login_container, show="*", width=20) 
password_entry.grid(row=3, column=0, padx=10, pady=10, columnspan=2)

login_button = tb.Button(login_container, text="ログイン", command=login)
login_button.grid(row=4, column=0, padx=10, pady=10, columnspan=2)

root.mainloop()